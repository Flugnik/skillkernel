"""
Tests for skills/limiter/exporter.py

All tests use tmp_path via the domain_root fixture from conftest.py.
No real project files are touched.
"""

from __future__ import annotations

import json
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from skills.limiter.domain import (
    OrderItemRecord,
    OrderRecord,
    OrderStatus,
    Product,
    ProductionDay,
    ProductionDayLimit,
)
from skills.limiter.exporter import build_export_bytes, export_path

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

DATE = date(2026, 3, 26)

PRODUCTS = [
    Product(sku="milk_1_5",  title="Молоко 1.5%",  price=65.0,  limit_default=100, unit="л"),
    Product(sku="kefir_1",   title="Кефир 1%",      price=70.0,  limit_default=60,  unit="л"),
    Product(sku="butter_72", title="Масло 72.5%",   price=350.0, limit_default=20,  unit="кг"),
]

DAY = ProductionDay(
    delivery_date=DATE,
    limits=[
        ProductionDayLimit(sku="milk_1_5",  limit=100),
        ProductionDayLimit(sku="kefir_1",   limit=60),
        ProductionDayLimit(sku="butter_72", limit=20),
    ],
)


def _make_order(
    order_id: str,
    status: OrderStatus,
    client_name: str | None,
    items: list[tuple[str, int]],
    note: str | None = None,
) -> OrderRecord:
    return OrderRecord(
        id=order_id,
        delivery_date=DATE,
        client_name=client_name,
        status=status,
        note=note,
        items=[
            OrderItemRecord(
                sku=sku,
                requested_qty=qty,
                accepted_qty=qty,
                price=65.0,
                line_total=qty * 65.0,
            )
            for sku, qty in items
        ],
    )


def _load_wb(xlsx_bytes: bytes) -> openpyxl.Workbook:
    from io import BytesIO
    return openpyxl.load_workbook(BytesIO(xlsx_bytes))


# ---------------------------------------------------------------------------
# 1. export creates xlsx bytes (non-empty)
# ---------------------------------------------------------------------------

def test_build_export_bytes_returns_bytes():
    orders = [_make_order("1", OrderStatus.Confirmed, "Иван", [("milk_1_5", 10)])]
    result = build_export_bytes(DATE, PRODUCTS, DAY, orders)
    assert isinstance(result, bytes)
    assert len(result) > 0


# ---------------------------------------------------------------------------
# 2. xlsx has exactly 2 sheets: Сводка and По клиентам
# ---------------------------------------------------------------------------

def test_two_sheets():
    orders = [_make_order("1", OrderStatus.Confirmed, "Иван", [("milk_1_5", 5)])]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    assert wb.sheetnames == ["Сводка", "По клиентам"]


# ---------------------------------------------------------------------------
# 3. Сводка contains correct limit / reserved / free
# ---------------------------------------------------------------------------

def test_summary_sheet_values():
    orders = [
        _make_order("1", OrderStatus.Confirmed, "Иван",  [("milk_1_5", 30), ("kefir_1", 10)]),
        _make_order("2", OrderStatus.Written,   "Мария", [("milk_1_5", 20)]),
    ]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["Сводка"]

    # Header row
    assert ws.cell(1, 1).value == "Товар"
    assert ws.cell(1, 2).value == "Цена"
    assert ws.cell(1, 3).value == "Лимит"
    assert ws.cell(1, 4).value == "Бронь"
    assert ws.cell(1, 5).value == "Свободно"
    assert ws.cell(1, 6).value == "Сумма"

    # milk_1_5 row (row 2): price=65, limit=100, reserved=50, free=50, sum=3250
    assert ws.cell(2, 1).value == "Молоко 1.5%"
    assert ws.cell(2, 2).value == 65.0
    assert ws.cell(2, 3).value == 100
    assert ws.cell(2, 4).value == 50
    assert ws.cell(2, 5).value == 50
    assert ws.cell(2, 6).value == 3250.0

    # kefir_1 row (row 3): price=70, limit=60, reserved=10, free=50, sum=700
    assert ws.cell(3, 1).value == "Кефир 1%"
    assert ws.cell(3, 2).value == 70.0
    assert ws.cell(3, 3).value == 60
    assert ws.cell(3, 4).value == 10
    assert ws.cell(3, 5).value == 50
    assert ws.cell(3, 6).value == 700.0

    # butter_72 row (row 4): price=350, limit=20, reserved=0, free=20, sum=0
    assert ws.cell(4, 1).value == "Масло 72.5%"
    assert ws.cell(4, 2).value == 350.0
    assert ws.cell(4, 3).value == 20
    assert ws.cell(4, 4).value == 0
    assert ws.cell(4, 5).value == 20
    assert ws.cell(4, 6).value == 0.0

    # Totals
    assert ws.cell(5, 1).value == "Итого по броне"
    assert ws.cell(5, 6).value == 3950.0
    assert ws.cell(6, 1).value == "Потенциал при полном лимите"
    assert ws.cell(6, 6).value == 17700.0


# ---------------------------------------------------------------------------
# 4. По клиентам has one row per order
# ---------------------------------------------------------------------------

def test_clients_sheet_one_row_per_order():
    orders = [
        _make_order("1", OrderStatus.Confirmed, "Иван",  [("milk_1_5", 5)]),
        _make_order("2", OrderStatus.Confirmed, "Мария", [("kefir_1",  3)]),
        _make_order("3", OrderStatus.Written,   "Петр",  [("butter_72", 2)]),
    ]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["По клиентам"]
    # 1 header + 3 orders = 4 rows
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    assert len(data_rows) == 3


# ---------------------------------------------------------------------------
# 5. Same client name → separate rows (not merged)
# ---------------------------------------------------------------------------

def test_same_client_not_merged():
    orders = [
        _make_order("1", OrderStatus.Confirmed, "Иван", [("milk_1_5", 5)]),
        _make_order("2", OrderStatus.Confirmed, "Иван", [("kefir_1",  3)]),
    ]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["По клиентам"]
    names = [ws.cell(row=r, column=1).value for r in range(2, 4)]
    assert names == ["Иван", "Иван"]


# ---------------------------------------------------------------------------
# 6. Empty client_name → "Без имени"
# ---------------------------------------------------------------------------

def test_empty_client_name_becomes_bez_imeni():
    orders = [_make_order("1", OrderStatus.Confirmed, None, [("milk_1_5", 5)])]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["По клиентам"]
    assert ws.cell(2, 1).value == "Без имени"


# ---------------------------------------------------------------------------
# 7. Missing product in order → empty cell
# ---------------------------------------------------------------------------

def test_missing_product_cell_is_empty():
    # Order only has milk_1_5; kefir_1 and butter_72 cells should be empty
    orders = [_make_order("1", OrderStatus.Confirmed, "Иван", [("milk_1_5", 7)])]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["По клиентам"]
    # col 1=ФИО, col 2=milk_1_5, col 3=kefir_1, col 4=butter_72, col 5=Примечание
    assert ws.cell(2, 2).value == 7    # milk present
    assert ws.cell(2, 3).value is None  # kefir absent
    assert ws.cell(2, 4).value is None  # butter absent


# ---------------------------------------------------------------------------
# 8. Cancelled / Draft orders are excluded
# ---------------------------------------------------------------------------

def test_cancelled_and_draft_excluded():
    orders = [
        _make_order("1", OrderStatus.Confirmed,          "Иван",  [("milk_1_5", 10)]),
        _make_order("2", OrderStatus.Cancelled,          "Отмена",[("milk_1_5", 99)]),
        _make_order("3", OrderStatus.Draft,              "Черн",  [("milk_1_5", 99)]),
        _make_order("4", OrderStatus.NeedsConfirmLimit,  "Нужно", [("milk_1_5", 99)]),
    ]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))

    # Сводка: only Иван's 10 should be reserved
    ws_sum = wb["Сводка"]
    assert ws_sum.cell(2, 4).value == 10  # milk reserved = 10

    # По клиентам: only 1 data row
    ws_cli = wb["По клиентам"]
    data_rows = [r for r in ws_cli.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    assert len(data_rows) == 1
    assert ws_cli.cell(2, 1).value == "Иван"


# ---------------------------------------------------------------------------
# 9. Product column order matches products list order
# ---------------------------------------------------------------------------

def test_product_column_order():
    orders = []
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["По клиентам"]
    # col 1=ФИО, col 2=Молоко 1.5%, col 3=Кефир 1%, col 4=Масло 72.5%, col 5=Примечание
    assert ws.cell(1, 1).value == "ФИО"
    assert ws.cell(1, 2).value == "Молоко 1.5%"
    assert ws.cell(1, 3).value == "Кефир 1%"
    assert ws.cell(1, 4).value == "Масло 72.5%"
    assert ws.cell(1, 5).value == "Примечание"


# ---------------------------------------------------------------------------
# 10. export_path returns correct path (uses manifest EXPORTS_DIR)
# ---------------------------------------------------------------------------

def test_export_path(domain_root: Path):
    path = export_path(DATE)
    assert path.endswith("limiter_export_2026-03-26.xlsx")
    assert "exports" in path


# ---------------------------------------------------------------------------
# 11. Executor writes file to disk (integration: bytes → file)
# ---------------------------------------------------------------------------

def test_executor_writes_file(domain_root: Path):
    """Executor reads domain data itself and writes a valid xlsx file."""
    import json
    from core.models import Action, ActionType
    from executors.file_executor import execute_write_xlsx_export

    # Persist a production day so the executor can load it
    day_data = DAY.model_dump(mode="json")
    day_path = domain_root / "production_days" / "2026-03-26.json"
    day_path.write_text(json.dumps(day_data, ensure_ascii=False), encoding="utf-8")

    out_path = domain_root / "exports" / "limiter_export_2026-03-26.xlsx"
    action = Action(
        action_type=ActionType.write_xlsx_export,
        params={"path": str(out_path), "delivery_date": DATE.isoformat()},
    )
    execute_write_xlsx_export(action)

    assert out_path.exists()
    assert out_path.stat().st_size > 0

    # Verify it's a valid xlsx
    wb = openpyxl.load_workbook(str(out_path))
    assert "Сводка" in wb.sheetnames
    assert "По клиентам" in wb.sheetnames


# ---------------------------------------------------------------------------
# 12. Note column is populated when order.note is set
# ---------------------------------------------------------------------------

def test_note_column_populated():
    orders = [_make_order("1", OrderStatus.Confirmed, "Иван", [("milk_1_5", 5)], note="Срочно")]
    wb = _load_wb(build_export_bytes(DATE, PRODUCTS, DAY, orders))
    ws = wb["По клиентам"]
    note_col = len(PRODUCTS) + 2  # 1(ФИО) + 3(products) + 1 = col 5
    assert ws.cell(2, note_col).value == "Срочно"


# ---------------------------------------------------------------------------
# 13. planner.build_export_plan returns plan_ready ActionPlan (domain_root fixture)
# ---------------------------------------------------------------------------

def test_build_export_plan_returns_action_plan(domain_root: Path):
    """build_export_plan stores only JSON-serialisable params (no bytes)."""
    from skills.limiter import planner
    from core.models import ActionType

    plan = planner.build_export_plan(event_id="test-evt", delivery_date=DATE)

    assert plan.skill_name == "limiter"
    assert len(plan.actions) == 1
    assert plan.actions[0].action_type == ActionType.write_xlsx_export

    params = plan.actions[0].params
    assert "delivery_date" in params
    assert params["delivery_date"] == DATE.isoformat()
    assert "path" in params
    assert "data_bytes" not in params
    assert "2026-03-26" in plan.preview_text

    # Must be JSON-serialisable (the key invariant)
    dumped = plan.model_dump(mode="json")
    assert dumped["actions"][0]["params"]["delivery_date"] == DATE.isoformat()
