"""
Limiter exporter — builds xlsx files for a given delivery date.

Two sheets:
  - Сводка:      Товар | Лимит | Забронировано | Свободно
  - По клиентам: ФИО | <product columns in products.json order> | Примечание

Only orders with status Confirmed or Written are included.
Draft / Cancelled / NeedsConfirmLimit are excluded.

No side effects here — returns the bytes of the workbook.
The caller (executor) is responsible for writing to disk.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font

from skills.limiter.domain import (
    OrderRecord,
    OrderStatus,
    Product,
    ProductionDay,
)

# Statuses included in the export
_EXPORT_STATUSES = {OrderStatus.Confirmed, OrderStatus.Written}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_export_bytes(
    delivery_date: date,
    products: list[Product],
    day: Optional[ProductionDay],
    orders: list[OrderRecord],
) -> bytes:
    """Build an xlsx workbook and return its raw bytes.

    Args:
        delivery_date: The date being exported.
        products:      Full ordered product list (defines column order).
        day:           ProductionDay for the date (may be None if not created).
        orders:        All orders for the date (will be filtered by status).

    Returns:
        Raw bytes of the xlsx file.
    """
    active_orders = [o for o in orders if o.status in _EXPORT_STATUSES]

    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, products, day, active_orders)
    _build_clients_sheet(wb, products, active_orders)

    # Remove the default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _build_summary_sheet(
    wb: openpyxl.Workbook,
    products: list[Product],
    day: Optional[ProductionDay],
    active_orders: list[OrderRecord],
) -> None:
    """Build the 'Сводка' sheet."""
    ws = wb.create_sheet("Сводка")

    # Header row
    headers = ["Товар", "Лимит", "Забронировано", "Свободно"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Build limit lookup from production day
    limit_by_sku: dict[str, int] = {}
    if day is not None:
        for lim in day.limits:
            limit_by_sku[lim.sku] = lim.limit

    # Build reserved lookup from active orders
    reserved_by_sku: dict[str, int] = {}
    for order in active_orders:
        for item in order.items:
            reserved_by_sku[item.sku] = reserved_by_sku.get(item.sku, 0) + item.accepted_qty

    # Data rows — in products.json order
    for row_idx, product in enumerate(products, start=2):
        limit = limit_by_sku.get(product.sku, 0)
        reserved = reserved_by_sku.get(product.sku, 0)
        free = limit - reserved
        ws.cell(row=row_idx, column=1, value=product.title)
        ws.cell(row=row_idx, column=2, value=limit)
        ws.cell(row=row_idx, column=3, value=reserved)
        ws.cell(row=row_idx, column=4, value=free)


def _build_clients_sheet(
    wb: openpyxl.Workbook,
    products: list[Product],
    active_orders: list[OrderRecord],
) -> None:
    """Build the 'По клиентам' sheet."""
    ws = wb.create_sheet("По клиентам")

    # Header row: ФИО | product titles... | Примечание
    headers = ["ФИО"] + [p.title for p in products] + ["Примечание"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Build a quick sku→column index (1-based, offset by 1 for ФИО column)
    sku_to_col: dict[str, int] = {
        p.sku: idx + 2  # col 1 = ФИО, col 2 = first product
        for idx, p in enumerate(products)
    }
    note_col = len(products) + 2  # last column

    # One row per order (do NOT merge same client)
    for row_idx, order in enumerate(active_orders, start=2):
        client_name = order.client_name or "Без имени"
        ws.cell(row=row_idx, column=1, value=client_name)

        for item in order.items:
            col = sku_to_col.get(item.sku)
            if col is not None:
                ws.cell(row=row_idx, column=col, value=item.accepted_qty)

        if order.note:
            ws.cell(row=row_idx, column=note_col, value=order.note)


# ---------------------------------------------------------------------------
# Path helper (mirrors manifest, but kept here for convenience)
# ---------------------------------------------------------------------------


def export_path(delivery_date: date) -> str:
    """Return the canonical export file path string for a given date."""
    import skills.limiter.manifest as _mf
    exports_dir = Path(_mf.EXPORTS_DIR)
    return str(exports_dir / f"limiter_export_{delivery_date.isoformat()}.xlsx")
